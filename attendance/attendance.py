import pandas as pd
import utils
from exceptions import *
from gui.frames.progress_frame import ProgressFrame
from .responses import Responses
from .rosters import Rosters
import openpyxl


# Attendance workbook
class Attendance:
    # TODO: Validate Excel file formatting
    def __init__(self,
                 responses_path: str,
                 rosters_path: str,
                 week: int,
                 progress_frame: ProgressFrame):
        self.week = week
        self.progress_frame = progress_frame

        try:
            self.responses = Responses(responses_path)
            self.rosters = Rosters(rosters_path)
            self.write_workbook = openpyxl.load_workbook(rosters_path)
        except ResponsesError as e:
            raise e
        except RostersError as e:
            raise e
        except OSError as e:
            raise e

        # Validate week
        if self.week < 1 or self.week > self.rosters.num_weeks:
            raise WeekRangeError(f'Week must be in the range [1, {self.rosters.num_weeks}]')

    def search_sheet_by_id(self, _id: str, course: str) -> int:
        sheet = self.rosters.workbook[course]
        row = sheet.loc[sheet['ID'] == _id]

        if not row.empty:
            print('SUCCESS')
            return row.index[0]
        print('FAILURE')
        return -1

    def search_sheet_by_last_name(self, last: str, course: str) -> int:
        sheet = self.rosters.workbook[course]
        row = sheet.loc[sheet['Name'].str.upper().str.startswith(last.upper())]

        if not row.empty:
            print('SUCCESS')
            return row.index[0]
        print('FAILURE')
        return -1

    def find_student_in_sheet(self, _id: str, last: str, course: str):
        # 1. Search by _id in current course sheet
        print('\tSearching by ID...', end=' ')
        row = self.search_sheet_by_id(_id, course)
        if row != -1:
            return row
        # 2. Search by last name in current course sheet
        print('\tSearching by Last Name...', end=' ')
        row = self.search_sheet_by_last_name(last, course)
        if row != -1:
            return row

    def increment_student(self, course: str, row: int, col: int) -> None:
        worksheet = self.write_workbook.get_sheet_by_name(course)
        val = worksheet.cell(row + 2, col).value
        worksheet.cell(row + 2, col).value = 1 if type(val) != int else val + 1

    def parse_response(self, response: pd.Series):
        date, last, _id, course = response
        course = utils.format_course(course)

        print(f'Searching for: {last, _id, course}')

        if course in self.rosters:
            print(f'\tCourse found in attendance workbook: {course}')
            row = self.find_student_in_sheet(_id, last, course)

            # Found student in current sheet
            if row is not None:
                # TODO: Validate responses match week
                # Ignore first 2 columns [ID, Name], indexing starts from 1
                col = 2 + self.week
                self.increment_student(course, row, col)

            # Unable to find student in sheet, search other sheets
            else:
                print('FAILURE: Unable to find student. Marking response.')

        print()
        self.progress_frame.update_progress_bar()

    def tally_responses(self) -> bool:
        self.responses.df.apply(self.parse_response, axis=1)
        self.write_workbook.save(self.rosters.path)   # Output to original file
        self.write_workbook.close()
        return True
